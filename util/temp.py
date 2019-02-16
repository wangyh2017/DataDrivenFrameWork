#encoding=utf-8
import openpyxl
from openpyxl.styles import Border,Side,Font
import time


from openpyxl import load_workbook

from openpyxl.styles import Border, Side, Font

import time



class parseExcel(object):

    def __init__(self, excelPath):
        self.excelPath = excelPath

        self.workbook = load_workbook(excelPath)  # 加载excel

        self.sheet = self.workbook.active  # 获取第一个sheet

        self.font = Font(color=None)

        self.colorDict = {"red": 'FFFF3030', "green": 'FF008B00'}

    # 设置当前要操作的sheet对象，使用index来获取相应的sheet

    def set_sheet_by_index(self, sheet_index):
        sheet_name = self.workbook.get_sheet_names()[sheet_index]

        self.sheet = self.workbook.get_sheet_by_name(sheet_name)

        return self.sheet

    # 获取当前默认sheet的名字

    def get_default_sheet(self):
        return self.sheet.title

    # 设置当前要操作的sheet对象，使用sheet名称来获取相应的sheet

    def set_sheet_by_name(self, sheet_name):
        sheet = self.workbook.get_sheet_by_name(sheet_name)

        self.sheet = sheet

        return self.sheet

    # 获取默认sheet中最大的行数

    def get_max_row_no(self):
        return self.sheet.max_row

    # 获取默认 sheet 的最大列数

    def get_max_col_no(self):
        return self.sheet.max_column

    # 获取默认sheet的最小（起始）行号

    def get_min_row_no(self):
        return self.sheet.min_row

    # 获取默认sheet的最小（起始）列号

    def get_min_col_no(self):
        return self.sheet.min_column

    # 获取默认 sheet 的所有行对象，

    def get_all_rows(self):
        return list(self.sheet.iter_rows())

        # return list(self.rows)也可以

    # 获取默认sheet中的所有列对象

    def get_all_cols(self):
        return list(self.sheet.iter_cols())

        # return list(self.sheet.columns)也可以

    # 从默认sheet中获取某一列，第一列从0开始

    def get_single_col(self, col_no):
        return self.get_all_cols()[col_no]

    # 从默认sheet中获取某一行，第一行从0开始

    def get_single_row(self, row_no):
        return self.get_all_rows()[row_no]

    # 从默认sheet中，通过行号和列号获取指定的单元格，注意行号和列号从1开始

    def get_cell(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no)

    # 从默认sheet中，通过行号和列号获取指定的单元格中的内容，注意行号和列号从1开始

    def get_cell_content(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定单元格中写入指定内容，注意行号和列号从1开始

    # 调用此方法的时候，excel不要处于打开状态

    def write_cell_content(self, row_no, col_no, content, font=None):
        self.sheet.cell(row=row_no, column=col_no).value = content

        self.workbook.save(self.excelPath)

        return self.sheet.cell(row=row_no, column=col_no).value

    # 从默认sheet中，通过行号和列号向指定单元格中写入当前日期，注意行号和列号从1开始

    # 调用此方法的时候，excel不要处于打开状态

    def write_cell_current_time(self, row_no, col_no):
        time1 = time.strftime("%Y-%m-%d %H:%M:%S")

        self.sheet.cell(row=row_no, column=col_no).value = str(time1)

        self.workbook.save(self.excelPath)

        return self.sheet.cell(row=row_no, column=col_no).value

    def save_excel_file(self):
        self.workbook.save(self.excelPath)


if __name__ == '__main__':
    # 测试代码

    p = parseExcel(u'/Users/wangyanhua/Documents/学习/Python/sele_frame/chapt_15/testData/126邮箱联系人3.xlsx')

    print u"获取默认sheet：", p.get_default_sheet()

    print u"设置sheet索引为1", p.set_sheet_by_index(1)

    print u"获取默认sheet：", p.get_default_sheet()

    print u"设置sheet索引为0", p.set_sheet_by_index(0)

    print u"获取默认sheet：", p.get_default_sheet()

    print u"最大行数：", p.get_max_row_no()

    print u"最大列数：", p.get_max_col_no()

    print u"最小起始行数：", p.get_min_row_no()

    print u"最小起始列数：", p.get_min_col_no()

    print u"所有行对象：", p.get_all_rows()

    print u"所有列对象：", p.get_all_cols()

    print u"获取某一列(2)：", p.get_single_col(2)

    print u"获取某一行(1)：", p.get_single_row(1)

    print u"取得行号和列号(2,2)单元格：", p.get_cell(2, 2)

    print u"取得行号和列号单元格的内容(2,2)", p.get_cell_content(2, 2)

    print u"行号和列号写入内容(11,11)：'xiaxiaoxu'", p.write_cell_content(11, 11, 'xiaxiaoxu')  #

    print u"行号和列号写入当前日期(13,13)：", p.write_cell_current_time(13, 13)
"""
获取默认sheet： 126账号
设置sheet索引为1 <Worksheet "\u8054\u7cfb\u4eba">
获取默认sheet： 联系人
设置sheet索引为0 <Worksheet "126\u8d26\u53f7">
获取默认sheet： 126账号
最大行数： 13
最大列数： 13
最小起始行数： 1
最小起始列数： 1
所有行对象： [(<Cell u'126\u8d26\u53f7'.A1>, <Cell u'126\u8d26\u53f7'.B1>, <Cell u'126\u8d26\u53f7'.C1>, <Cell u'126\u8d26\u53f7'.D1>, <Cell u'126\u8d26\u53f7'.E1>, <Cell u'126\u8d26\u53f7'.F1>, <Cell u'126\u8d26\u53f7'.G1>, <Cell u'126\u8d26\u53f7'.H1>, <Cell u'126\u8d26\u53f7'.I1>, <Cell u'126\u8d26\u53f7'.J1>, <Cell u'126\u8d26\u53f7'.K1>, <Cell u'126\u8d26\u53f7'.L1>, <Cell u'126\u8d26\u53f7'.M1>), (<Cell u'126\u8d26\u53f7'.A2>, <Cell u'126\u8d26\u53f7'.B2>, <Cell u'126\u8d26\u53f7'.C2>, <Cell u'126\u8d26\u53f7'.D2>, <Cell u'126\u8d26\u53f7'.E2>, <Cell u'126\u8d26\u53f7'.F2>, <Cell u'126\u8d26\u53f7'.G2>, <Cell u'126\u8d26\u53f7'.H2>, <Cell u'126\u8d26\u53f7'.I2>, <Cell u'126\u8d26\u53f7'.J2>, <Cell u'126\u8d26\u53f7'.K2>, <Cell u'126\u8d26\u53f7'.L2>, <Cell u'126\u8d26\u53f7'.M2>), (<Cell u'126\u8d26\u53f7'.A3>, <Cell u'126\u8d26\u53f7'.B3>, <Cell u'126\u8d26\u53f7'.C3>, <Cell u'126\u8d26\u53f7'.D3>, <Cell u'126\u8d26\u53f7'.E3>, <Cell u'126\u8d26\u53f7'.F3>, <Cell u'126\u8d26\u53f7'.G3>, <Cell u'126\u8d26\u53f7'.H3>, <Cell u'126\u8d26\u53f7'.I3>, <Cell u'126\u8d26\u53f7'.J3>, <Cell u'126\u8d26\u53f7'.K3>, <Cell u'126\u8d26\u53f7'.L3>, <Cell u'126\u8d26\u53f7'.M3>), (<Cell u'126\u8d26\u53f7'.A4>, <Cell u'126\u8d26\u53f7'.B4>, <Cell u'126\u8d26\u53f7'.C4>, <Cell u'126\u8d26\u53f7'.D4>, <Cell u'126\u8d26\u53f7'.E4>, <Cell u'126\u8d26\u53f7'.F4>, <Cell u'126\u8d26\u53f7'.G4>, <Cell u'126\u8d26\u53f7'.H4>, <Cell u'126\u8d26\u53f7'.I4>, <Cell u'126\u8d26\u53f7'.J4>, <Cell u'126\u8d26\u53f7'.K4>, <Cell u'126\u8d26\u53f7'.L4>, <Cell u'126\u8d26\u53f7'.M4>), (<Cell u'126\u8d26\u53f7'.A5>, <Cell u'126\u8d26\u53f7'.B5>, <Cell u'126\u8d26\u53f7'.C5>, <Cell u'126\u8d26\u53f7'.D5>, <Cell u'126\u8d26\u53f7'.E5>, <Cell u'126\u8d26\u53f7'.F5>, <Cell u'126\u8d26\u53f7'.G5>, <Cell u'126\u8d26\u53f7'.H5>, <Cell u'126\u8d26\u53f7'.I5>, <Cell u'126\u8d26\u53f7'.J5>, <Cell u'126\u8d26\u53f7'.K5>, <Cell u'126\u8d26\u53f7'.L5>, <Cell u'126\u8d26\u53f7'.M5>), (<Cell u'126\u8d26\u53f7'.A6>, <Cell u'126\u8d26\u53f7'.B6>, <Cell u'126\u8d26\u53f7'.C6>, <Cell u'126\u8d26\u53f7'.D6>, <Cell u'126\u8d26\u53f7'.E6>, <Cell u'126\u8d26\u53f7'.F6>, <Cell u'126\u8d26\u53f7'.G6>, <Cell u'126\u8d26\u53f7'.H6>, <Cell u'126\u8d26\u53f7'.I6>, <Cell u'126\u8d26\u53f7'.J6>, <Cell u'126\u8d26\u53f7'.K6>, <Cell u'126\u8d26\u53f7'.L6>, <Cell u'126\u8d26\u53f7'.M6>), (<Cell u'126\u8d26\u53f7'.A7>, <Cell u'126\u8d26\u53f7'.B7>, <Cell u'126\u8d26\u53f7'.C7>, <Cell u'126\u8d26\u53f7'.D7>, <Cell u'126\u8d26\u53f7'.E7>, <Cell u'126\u8d26\u53f7'.F7>, <Cell u'126\u8d26\u53f7'.G7>, <Cell u'126\u8d26\u53f7'.H7>, <Cell u'126\u8d26\u53f7'.I7>, <Cell u'126\u8d26\u53f7'.J7>, <Cell u'126\u8d26\u53f7'.K7>, <Cell u'126\u8d26\u53f7'.L7>, <Cell u'126\u8d26\u53f7'.M7>), (<Cell u'126\u8d26\u53f7'.A8>, <Cell u'126\u8d26\u53f7'.B8>, <Cell u'126\u8d26\u53f7'.C8>, <Cell u'126\u8d26\u53f7'.D8>, <Cell u'126\u8d26\u53f7'.E8>, <Cell u'126\u8d26\u53f7'.F8>, <Cell u'126\u8d26\u53f7'.G8>, <Cell u'126\u8d26\u53f7'.H8>, <Cell u'126\u8d26\u53f7'.I8>, <Cell u'126\u8d26\u53f7'.J8>, <Cell u'126\u8d26\u53f7'.K8>, <Cell u'126\u8d26\u53f7'.L8>, <Cell u'126\u8d26\u53f7'.M8>), (<Cell u'126\u8d26\u53f7'.A9>, <Cell u'126\u8d26\u53f7'.B9>, <Cell u'126\u8d26\u53f7'.C9>, <Cell u'126\u8d26\u53f7'.D9>, <Cell u'126\u8d26\u53f7'.E9>, <Cell u'126\u8d26\u53f7'.F9>, <Cell u'126\u8d26\u53f7'.G9>, <Cell u'126\u8d26\u53f7'.H9>, <Cell u'126\u8d26\u53f7'.I9>, <Cell u'126\u8d26\u53f7'.J9>, <Cell u'126\u8d26\u53f7'.K9>, <Cell u'126\u8d26\u53f7'.L9>, <Cell u'126\u8d26\u53f7'.M9>), (<Cell u'126\u8d26\u53f7'.A10>, <Cell u'126\u8d26\u53f7'.B10>, <Cell u'126\u8d26\u53f7'.C10>, <Cell u'126\u8d26\u53f7'.D10>, <Cell u'126\u8d26\u53f7'.E10>, <Cell u'126\u8d26\u53f7'.F10>, <Cell u'126\u8d26\u53f7'.G10>, <Cell u'126\u8d26\u53f7'.H10>, <Cell u'126\u8d26\u53f7'.I10>, <Cell u'126\u8d26\u53f7'.J10>, <Cell u'126\u8d26\u53f7'.K10>, <Cell u'126\u8d26\u53f7'.L10>, <Cell u'126\u8d26\u53f7'.M10>), (<Cell u'126\u8d26\u53f7'.A11>, <Cell u'126\u8d26\u53f7'.B11>, <Cell u'126\u8d26\u53f7'.C11>, <Cell u'126\u8d26\u53f7'.D11>, <Cell u'126\u8d26\u53f7'.E11>, <Cell u'126\u8d26\u53f7'.F11>, <Cell u'126\u8d26\u53f7'.G11>, <Cell u'126\u8d26\u53f7'.H11>, <Cell u'126\u8d26\u53f7'.I11>, <Cell u'126\u8d26\u53f7'.J11>, <Cell u'126\u8d26\u53f7'.K11>, <Cell u'126\u8d26\u53f7'.L11>, <Cell u'126\u8d26\u53f7'.M11>), (<Cell u'126\u8d26\u53f7'.A12>, <Cell u'126\u8d26\u53f7'.B12>, <Cell u'126\u8d26\u53f7'.C12>, <Cell u'126\u8d26\u53f7'.D12>, <Cell u'126\u8d26\u53f7'.E12>, <Cell u'126\u8d26\u53f7'.F12>, <Cell u'126\u8d26\u53f7'.G12>, <Cell u'126\u8d26\u53f7'.H12>, <Cell u'126\u8d26\u53f7'.I12>, <Cell u'126\u8d26\u53f7'.J12>, <Cell u'126\u8d26\u53f7'.K12>, <Cell u'126\u8d26\u53f7'.L12>, <Cell u'126\u8d26\u53f7'.M12>), (<Cell u'126\u8d26\u53f7'.A13>, <Cell u'126\u8d26\u53f7'.B13>, <Cell u'126\u8d26\u53f7'.C13>, <Cell u'126\u8d26\u53f7'.D13>, <Cell u'126\u8d26\u53f7'.E13>, <Cell u'126\u8d26\u53f7'.F13>, <Cell u'126\u8d26\u53f7'.G13>, <Cell u'126\u8d26\u53f7'.H13>, <Cell u'126\u8d26\u53f7'.I13>, <Cell u'126\u8d26\u53f7'.J13>, <Cell u'126\u8d26\u53f7'.K13>, <Cell u'126\u8d26\u53f7'.L13>, <Cell u'126\u8d26\u53f7'.M13>)]
所有列对象： [(<Cell u'126\u8d26\u53f7'.A1>, <Cell u'126\u8d26\u53f7'.A2>, <Cell u'126\u8d26\u53f7'.A3>, <Cell u'126\u8d26\u53f7'.A4>, <Cell u'126\u8d26\u53f7'.A5>, <Cell u'126\u8d26\u53f7'.A6>, <Cell u'126\u8d26\u53f7'.A7>, <Cell u'126\u8d26\u53f7'.A8>, <Cell u'126\u8d26\u53f7'.A9>, <Cell u'126\u8d26\u53f7'.A10>, <Cell u'126\u8d26\u53f7'.A11>, <Cell u'126\u8d26\u53f7'.A12>, <Cell u'126\u8d26\u53f7'.A13>), (<Cell u'126\u8d26\u53f7'.B1>, <Cell u'126\u8d26\u53f7'.B2>, <Cell u'126\u8d26\u53f7'.B3>, <Cell u'126\u8d26\u53f7'.B4>, <Cell u'126\u8d26\u53f7'.B5>, <Cell u'126\u8d26\u53f7'.B6>, <Cell u'126\u8d26\u53f7'.B7>, <Cell u'126\u8d26\u53f7'.B8>, <Cell u'126\u8d26\u53f7'.B9>, <Cell u'126\u8d26\u53f7'.B10>, <Cell u'126\u8d26\u53f7'.B11>, <Cell u'126\u8d26\u53f7'.B12>, <Cell u'126\u8d26\u53f7'.B13>), (<Cell u'126\u8d26\u53f7'.C1>, <Cell u'126\u8d26\u53f7'.C2>, <Cell u'126\u8d26\u53f7'.C3>, <Cell u'126\u8d26\u53f7'.C4>, <Cell u'126\u8d26\u53f7'.C5>, <Cell u'126\u8d26\u53f7'.C6>, <Cell u'126\u8d26\u53f7'.C7>, <Cell u'126\u8d26\u53f7'.C8>, <Cell u'126\u8d26\u53f7'.C9>, <Cell u'126\u8d26\u53f7'.C10>, <Cell u'126\u8d26\u53f7'.C11>, <Cell u'126\u8d26\u53f7'.C12>, <Cell u'126\u8d26\u53f7'.C13>), (<Cell u'126\u8d26\u53f7'.D1>, <Cell u'126\u8d26\u53f7'.D2>, <Cell u'126\u8d26\u53f7'.D3>, <Cell u'126\u8d26\u53f7'.D4>, <Cell u'126\u8d26\u53f7'.D5>, <Cell u'126\u8d26\u53f7'.D6>, <Cell u'126\u8d26\u53f7'.D7>, <Cell u'126\u8d26\u53f7'.D8>, <Cell u'126\u8d26\u53f7'.D9>, <Cell u'126\u8d26\u53f7'.D10>, <Cell u'126\u8d26\u53f7'.D11>, <Cell u'126\u8d26\u53f7'.D12>, <Cell u'126\u8d26\u53f7'.D13>), (<Cell u'126\u8d26\u53f7'.E1>, <Cell u'126\u8d26\u53f7'.E2>, <Cell u'126\u8d26\u53f7'.E3>, <Cell u'126\u8d26\u53f7'.E4>, <Cell u'126\u8d26\u53f7'.E5>, <Cell u'126\u8d26\u53f7'.E6>, <Cell u'126\u8d26\u53f7'.E7>, <Cell u'126\u8d26\u53f7'.E8>, <Cell u'126\u8d26\u53f7'.E9>, <Cell u'126\u8d26\u53f7'.E10>, <Cell u'126\u8d26\u53f7'.E11>, <Cell u'126\u8d26\u53f7'.E12>, <Cell u'126\u8d26\u53f7'.E13>), (<Cell u'126\u8d26\u53f7'.F1>, <Cell u'126\u8d26\u53f7'.F2>, <Cell u'126\u8d26\u53f7'.F3>, <Cell u'126\u8d26\u53f7'.F4>, <Cell u'126\u8d26\u53f7'.F5>, <Cell u'126\u8d26\u53f7'.F6>, <Cell u'126\u8d26\u53f7'.F7>, <Cell u'126\u8d26\u53f7'.F8>, <Cell u'126\u8d26\u53f7'.F9>, <Cell u'126\u8d26\u53f7'.F10>, <Cell u'126\u8d26\u53f7'.F11>, <Cell u'126\u8d26\u53f7'.F12>, <Cell u'126\u8d26\u53f7'.F13>), (<Cell u'126\u8d26\u53f7'.G1>, <Cell u'126\u8d26\u53f7'.G2>, <Cell u'126\u8d26\u53f7'.G3>, <Cell u'126\u8d26\u53f7'.G4>, <Cell u'126\u8d26\u53f7'.G5>, <Cell u'126\u8d26\u53f7'.G6>, <Cell u'126\u8d26\u53f7'.G7>, <Cell u'126\u8d26\u53f7'.G8>, <Cell u'126\u8d26\u53f7'.G9>, <Cell u'126\u8d26\u53f7'.G10>, <Cell u'126\u8d26\u53f7'.G11>, <Cell u'126\u8d26\u53f7'.G12>, <Cell u'126\u8d26\u53f7'.G13>), (<Cell u'126\u8d26\u53f7'.H1>, <Cell u'126\u8d26\u53f7'.H2>, <Cell u'126\u8d26\u53f7'.H3>, <Cell u'126\u8d26\u53f7'.H4>, <Cell u'126\u8d26\u53f7'.H5>, <Cell u'126\u8d26\u53f7'.H6>, <Cell u'126\u8d26\u53f7'.H7>, <Cell u'126\u8d26\u53f7'.H8>, <Cell u'126\u8d26\u53f7'.H9>, <Cell u'126\u8d26\u53f7'.H10>, <Cell u'126\u8d26\u53f7'.H11>, <Cell u'126\u8d26\u53f7'.H12>, <Cell u'126\u8d26\u53f7'.H13>), (<Cell u'126\u8d26\u53f7'.I1>, <Cell u'126\u8d26\u53f7'.I2>, <Cell u'126\u8d26\u53f7'.I3>, <Cell u'126\u8d26\u53f7'.I4>, <Cell u'126\u8d26\u53f7'.I5>, <Cell u'126\u8d26\u53f7'.I6>, <Cell u'126\u8d26\u53f7'.I7>, <Cell u'126\u8d26\u53f7'.I8>, <Cell u'126\u8d26\u53f7'.I9>, <Cell u'126\u8d26\u53f7'.I10>, <Cell u'126\u8d26\u53f7'.I11>, <Cell u'126\u8d26\u53f7'.I12>, <Cell u'126\u8d26\u53f7'.I13>), (<Cell u'126\u8d26\u53f7'.J1>, <Cell u'126\u8d26\u53f7'.J2>, <Cell u'126\u8d26\u53f7'.J3>, <Cell u'126\u8d26\u53f7'.J4>, <Cell u'126\u8d26\u53f7'.J5>, <Cell u'126\u8d26\u53f7'.J6>, <Cell u'126\u8d26\u53f7'.J7>, <Cell u'126\u8d26\u53f7'.J8>, <Cell u'126\u8d26\u53f7'.J9>, <Cell u'126\u8d26\u53f7'.J10>, <Cell u'126\u8d26\u53f7'.J11>, <Cell u'126\u8d26\u53f7'.J12>, <Cell u'126\u8d26\u53f7'.J13>), (<Cell u'126\u8d26\u53f7'.K1>, <Cell u'126\u8d26\u53f7'.K2>, <Cell u'126\u8d26\u53f7'.K3>, <Cell u'126\u8d26\u53f7'.K4>, <Cell u'126\u8d26\u53f7'.K5>, <Cell u'126\u8d26\u53f7'.K6>, <Cell u'126\u8d26\u53f7'.K7>, <Cell u'126\u8d26\u53f7'.K8>, <Cell u'126\u8d26\u53f7'.K9>, <Cell u'126\u8d26\u53f7'.K10>, <Cell u'126\u8d26\u53f7'.K11>, <Cell u'126\u8d26\u53f7'.K12>, <Cell u'126\u8d26\u53f7'.K13>), (<Cell u'126\u8d26\u53f7'.L1>, <Cell u'126\u8d26\u53f7'.L2>, <Cell u'126\u8d26\u53f7'.L3>, <Cell u'126\u8d26\u53f7'.L4>, <Cell u'126\u8d26\u53f7'.L5>, <Cell u'126\u8d26\u53f7'.L6>, <Cell u'126\u8d26\u53f7'.L7>, <Cell u'126\u8d26\u53f7'.L8>, <Cell u'126\u8d26\u53f7'.L9>, <Cell u'126\u8d26\u53f7'.L10>, <Cell u'126\u8d26\u53f7'.L11>, <Cell u'126\u8d26\u53f7'.L12>, <Cell u'126\u8d26\u53f7'.L13>), (<Cell u'126\u8d26\u53f7'.M1>, <Cell u'126\u8d26\u53f7'.M2>, <Cell u'126\u8d26\u53f7'.M3>, <Cell u'126\u8d26\u53f7'.M4>, <Cell u'126\u8d26\u53f7'.M5>, <Cell u'126\u8d26\u53f7'.M6>, <Cell u'126\u8d26\u53f7'.M7>, <Cell u'126\u8d26\u53f7'.M8>, <Cell u'126\u8d26\u53f7'.M9>, <Cell u'126\u8d26\u53f7'.M10>, <Cell u'126\u8d26\u53f7'.M11>, <Cell u'126\u8d26\u53f7'.M12>, <Cell u'126\u8d26\u53f7'.M13>)]
获取某一列(2)： (<Cell u'126\u8d26\u53f7'.C1>, <Cell u'126\u8d26\u53f7'.C2>, <Cell u'126\u8d26\u53f7'.C3>, <Cell u'126\u8d26\u53f7'.C4>, <Cell u'126\u8d26\u53f7'.C5>, <Cell u'126\u8d26\u53f7'.C6>, <Cell u'126\u8d26\u53f7'.C7>, <Cell u'126\u8d26\u53f7'.C8>, <Cell u'126\u8d26\u53f7'.C9>, <Cell u'126\u8d26\u53f7'.C10>, <Cell u'126\u8d26\u53f7'.C11>, <Cell u'126\u8d26\u53f7'.C12>, <Cell u'126\u8d26\u53f7'.C13>)
获取某一行(1)： (<Cell u'126\u8d26\u53f7'.A2>, <Cell u'126\u8d26\u53f7'.B2>, <Cell u'126\u8d26\u53f7'.C2>, <Cell u'126\u8d26\u53f7'.D2>, <Cell u'126\u8d26\u53f7'.E2>, <Cell u'126\u8d26\u53f7'.F2>, <Cell u'126\u8d26\u53f7'.G2>, <Cell u'126\u8d26\u53f7'.H2>, <Cell u'126\u8d26\u53f7'.I2>, <Cell u'126\u8d26\u53f7'.J2>, <Cell u'126\u8d26\u53f7'.K2>, <Cell u'126\u8d26\u53f7'.L2>, <Cell u'126\u8d26\u53f7'.M2>)
取得行号和列号(2,2)单元格： <Cell u'126\u8d26\u53f7'.B2>
取得行号和列号单元格的内容(2,2) xiaxiaoxu1987
行号和列号写入内容(11,11)：'xiaxiaoxu' xiaxiaoxu
行号和列号写入当前日期(13,13)： 2019-01-27 13:23:05



"""